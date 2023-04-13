import requests
import json
from bs4 import BeautifulSoup
from DAO import getCursor
import openpyxl

if __name__ == '__main__':
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.69",
        "Cookie": "cnfunco=1; cors_js=1; bkng_sso_session=e30; bkng_sso_ses=e30; _pxvid=a9278cf5-c273-11ed-abc2-61766a736e6a; aliyungf_tc=d3885ada5956aae4854d7f2d90a5450565aebdd37dba3f5cc48c6e44aa587d12; pxcts=1eaefc4b-c2e6-11ed-be83-6e44527a546c; _px3=b00ae55e3e40e14b8459a531f519a748224447082cb9522b158167ac90fcbc32:ycsVtiW3puL9qwm64tJrDhRLcMmNu37y58fRzDnsNhsjkJU1Yv2p6xd38D6fqbKcCRPFqc44cMKRkt/uc8SlVA==:1000:58kZIf28YEBeDrApusYwPF7y+MsyMeDTAFiahHCcQE83YEgoH5SL1V0GmyAaQtWI9f4lijGxnRT3NXG8vFOEqs/0RwLmBGuBxAqwpkKJdr3Eh6njLy4IzMPvDfKS/X7MFiJkQ4Ixc6baLIh5GeqFyXjFenr1nRAwnG5YbbZ1Gw4nXGOIWUdYklIOwYCezOYThUmkkXGCUIbHaDKxXMx/YA==; _pxde=e350cb609d3e9380ec12b46d348cda305235ff04a281c24c4c35d03b9703feab:eyJ0aW1lc3RhbXAiOjE2Nzg4NjA5MTYwNzMsImZfa2IiOjAsImlwY19pZCI6W119; acw_tc=781bad4d16788610098562653e3e457a574af3ef4a11e0de50339346fc8e86; OptanonConsent=isGpcEnabled=0&datestamp=Wed+Mar+15+2023+14%3A16%3A52+GMT%2B0800+(%E4%B8%AD%E5%9B%BD%E6%A0%87%E5%87%86%E6%97%B6%E9%97%B4)&version=6.22.0&isIABGlobal=false&hosts=&consentId=9ed25a52-ac48-498a-9154-02a2277c150b&interactionCount=1&landingPath=NotLandingPage&AwaitingReconsent=false&groups=C0001%3A1%2CC0002%3A0%2CC0004%3A0&implicitConsentCountry=GDPR&implicitConsentDate=1678803715087; bkng=11UmFuZG9tSVYkc2RlIyh9Yaa29%2F3xUOLbiKbS0JOgDBKd%2F%2B1hfoh7rsSN%2BxB95HtADwNYBs8pGf2lSb2tl%2FzFOhcqAGvO9287BuvqnpxJp%2FYvSeB0JnGKuKXONr6PaKaTqjIKtIFNTFi0c07sxHNaiLMvvFr3De4Csoj82Qf60AK8434Pk8LrQZQKiiVeiU4idAcAQtO2Reo%3D"
    }
    num = 0
    elem_list = []

    # 建立数据库链连接
    con, cur = getCursor()
    for offset in range(0, 10000, 25):
        # 构造url
        url = f"https://www.booking.cn/searchresults.zh-cn.html?label=gen173nr-1FCAEoggI46AdIM1gEaDGIAQGYASu4ARfIAQzYAQHoAQH4AQuIAgGoAgO4AtrP3aEGwAIB0gIkMGU3OTdmY2EtM2E1NC00ZTEwLWFmOGUtYWUxMTgxNGFjNmI42AIG4AIB&sid=497dbec8b193760b4d4de17e00e3c3f8&aid=304142&dest_id=44&dest_type=country&srpvid=a128113547a4018d&order=score&offset={offset}"

        # 发送请求，接收响应
        data = requests.get(url, headers=header).text

        # 解析html
        soup = BeautifulSoup(data, "html.parser")
        soup.prettify()

        # 获取你想要的数据
        cards = soup.find_all("div", attrs={"data-testid": "property-card"})
        for card in cards:
            try:
                # 酒店名
                name = card.find_next("div", attrs={"data-testid": "title"}).text
                flag = 0
                for el in elem_list:
                    if name in el.values():
                        flag = 1
                        print(f"第{offset}页{name}出现重复")
                if flag == 1:
                    break
                else:
                    elem = {"name":name}

                elem["id"] = num

                # 酒店链接
                link = card.find_next("a", attrs={"data-testid": "title-link"}).get('href')
                elem["link"] = link

                # 酒店简介
                description = card.find_next("div", attrs={"class": "d8eab2cf7f"}).text
                elem["description"] = description

                # 进入酒店链接获取更多信息
                info = requests.get(link, headers=header).text
                info_soup = BeautifulSoup(info, "html.parser")
                info_soup.prettify()

                # 所在地区
                positions = info_soup.find_all("a", attrs={"itemprop": "item"})
                country = positions[-4].text.replace("\n", "")
                if country != "中国":
                    country = positions[-5].text.replace("\n", "")
                    province = positions[-4].text.replace("\n", "")
                    city = positions[-3].text.replace("\n", "")
                    region = positions[-2].text.replace("\n", "")
                    area = country + ' ' + province + ' ' + city + ' ' + region
                    elem["area"] = area
                    elem["country"] = country
                    elem["province"] = province
                    elem["city"] = city
                    elem["region"] = region
                else:
                    province = positions[-3].text.replace("\n", "")
                    city = positions[-2].text.replace("\n", "")
                    area = country + ' ' + province + ' ' + city
                    region = ""
                    elem["area"] = area
                    elem["country"] = country
                    elem["province"] = province
                    elem["city"] = city
                    elem["region"] = region

                # 具体位置
                location = info_soup.find_all("span", attrs={"data-component": "tooltip"})[0].text.replace("\n", "")
                elem["location"] = location

                # 图片
                images_link = info_soup.find("div", attrs={
                    "class": "clearfix bh-photo-grid bh-photo-grid--space-down fix-score-hover-opacity"})
                images_link = images_link.find_all("div", attrs={"aria-hidden": "true"})
                for i in range(1, 4):
                    image_link = images_link[i].find_next("img").get("src").replace("\n", "")
                    im_res = requests.get(image_link)
                    with open("./images/%s %d.jpg" % (name, i), 'wb') as f:
                        f.write(im_res.content)
                    elem[f"image{i}"] = image_link

                # 读取图片并写入数据库
                images = []
                for i in range(1, 4):
                    with open("./images/%s %d.jpg" % (name, i), 'rb') as f:
                        images.append(f.read())
                stm = "INSERT INTO HOTELS(ID, LINK, NAME, DESCRIPTION, AREA, COUNTRY, PROVINCE, CITY, REGION, " \
                      "LOCATION, IMAGE1, IMAGE2, IMAGE3) " \
                      "VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                values = (num, link, name, description, area, country, province, city, region, location, images[0], images[1], images[2])
                cur.execute(stm, values)
                con.commit()
                print(elem)
                elem_list.append(elem)
                num += 1
            except Exception as e:
                print(e)
            finally:
                continue

    # 提交请求，关闭数据库
    con.commit()
    con.close()
    # 保存数据
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    sheet_header = ["id", "link", "name", "description", "area", "country", "province", "city", "location", "image1", "image2", "image3"]

    # 写入表头
    for col, h in enumerate(sheet_header, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = h

    # 写入数据
    for row, rd in enumerate(elem_list,2):
        for col, cd in rd.items():
            cell = worksheet.cell(row=row, column=sheet_header.index(col)+1)
            cell.value = cd

    # 保存到xlsx
    workbook.save(filename="data.xlsx")

    with open("data.json", "w") as f:
        f.write(json.dumps(elem_list, indent=4, separators=(',', ': ')))
